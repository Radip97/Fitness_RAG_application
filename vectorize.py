import os
import sys
import json
import time
import shutil
import logging
import argparse
import openpyxl
from pathlib import Path
from langchain_core.documents import Document
from langchain_text_splitters import RecursiveCharacterTextSplitter
from langchain_community.document_loaders import PyPDFLoader
from langchain_community.vectorstores import Chroma
from langchain_huggingface import HuggingFaceEmbeddings

# Import configurations from Fitness_App to maintain consistency
from Fitness_App import DATA_DIR, CHROMA_DIR, EMBEDDING_MODEL, CHUNK_SIZE, CHUNK_OVERLAP, CSV_ROWS_PER_DOC

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s | %(levelname)s | %(message)s",
)
log = logging.getLogger(__name__)

# ==============================================================
# Document Loaders (Moved from Fitness_App)
# ==============================================================

def load_pdfs(data_dir):
    """Load all PDF files using PyPDFLoader."""
    docs = []
    for pdf_file in sorted(data_dir.glob("*.pdf")):
        log.info(f"  Loading PDF: {pdf_file.name}")
        try:
            loader = PyPDFLoader(str(pdf_file))
            pages = loader.load()
            for page in pages:
                page.metadata["source"] = pdf_file.name
            docs.extend(pages)
            log.info(f"     -> {len(pages)} pages extracted")
        except Exception as e:
            log.warning(f"     Failed to load {pdf_file.name}: {e}")
    return docs


def load_csvs(data_dir):
    """Load CSV files, selecting key columns and grouping rows for efficiency."""
    import csv as csv_module
    KEY_COLS = {
        "Age", "Gender", "Weight (kg)", "Height (m)", "Max_BPM", "Avg_BPM",
        "Session_Duration (hours)", "Calories_Burned", "Workout_Type",
        "Fat_Percentage", "Water_Intake (liters)", "Workout_Frequency (days/week)",
        "Experience_Level", "BMI", "Daily meals frequency",
        "Carbs", "Proteins", "Fats", "Calories",
        "meal_name", "meal_type", "diet_type",
        "Name of Exercise", "Sets", "Reps", "Benefit",
        "Target Muscle Group", "Equipment Needed", "Difficulty Level",
        "Body Part", "Workout",
    }
    docs = []
    for csv_file in sorted(data_dir.glob("*.csv")):
        log.info(f"  Loading CSV: {csv_file.name}")
        try:
            with open(csv_file, "r", encoding="utf-8") as f:
                reader = csv_module.DictReader(f)
                headers = reader.fieldnames or []
                use_filter = len(headers) > 30
                if use_filter:
                    keep = [h for h in headers if h in KEY_COLS]
                    log.info(f"     Filtering {len(headers)} cols -> {len(keep)} key cols")
                else:
                    keep = headers
                batch = []
                row_count = 0
                for row in reader:
                    line = ", ".join(f"{k}: {row[k]}" for k in keep if row.get(k) and row[k].strip())
                    if line: batch.append(line)
                    row_count += 1
                    if len(batch) >= CSV_ROWS_PER_DOC:
                        docs.append(Document(page_content="\n".join(batch), metadata={"source": csv_file.name}))
                        batch = []
                if batch:
                    docs.append(Document(page_content="\n".join(batch), metadata={"source": csv_file.name}))
            log.info(f"     -> {row_count} rows grouped into documents")
        except Exception as e:
            log.warning(f"     Failed to load {csv_file.name}: {e}")
    return docs


def load_json_exercises(data_dir):
    """Load JSON exercise files and convert structured data to readable text."""
    docs = []
    for json_file in sorted(data_dir.glob("*.json")):
        log.info(f"  Loading JSON: {json_file.name}")
        try:
            with open(json_file, "r", encoding="utf-8") as f:
                exercises = json.load(f)
            for ex in exercises:
                name = ex.get("name", "Unknown Exercise")
                parts = [f"Exercise: {name}"]
                if "instructions" in ex and ex["instructions"]:
                    parts.append("Instructions: " + " ".join(ex["instructions"]))
                if "steps" in ex and ex["steps"]:
                    steps_text = " ".join(f"Step {i+1}: {s}" for i, s in enumerate(ex["steps"]))
                    parts.append(steps_text)
                if "primaryMuscles" in ex and ex["primaryMuscles"]:
                    parts.append(f"Primary muscles: {', '.join(ex['primaryMuscles'])}")
                if "secondaryMuscles" in ex and ex["secondaryMuscles"]:
                    parts.append(f"Secondary muscles: {', '.join(ex['secondaryMuscles'])}")
                if "notes" in ex and ex["notes"]:
                    parts.append(f"Notes: {ex['notes']}")
                text = "\n".join(parts)
                docs.append(Document(page_content=text, metadata={"source": json_file.name, "exercise_name": name}))
            log.info(f"     -> {len(exercises)} exercises loaded")
        except Exception as e:
            log.warning(f"     Failed to load {json_file.name}: {e}")
    return docs


def load_xlsx(data_dir):
    """Load XLSX files using openpyxl, converting each row to a Document."""
    docs = []
    for xlsx_file in sorted(data_dir.glob("*.xlsx")):
        log.info(f"  Loading XLSX: {xlsx_file.name}")
        try:
            wb = openpyxl.load_workbook(str(xlsx_file), read_only=True, data_only=True)
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                rows = list(ws.iter_rows(values_only=True))
                if not rows: continue
                headers = [str(h).strip() if h else f"col_{i}" for i, h in enumerate(rows[0])]
                count = 0
                for row in rows[1:]:
                    row_data = {}
                    for i, val in enumerate(row):
                        if i < len(headers) and val is not None:
                            row_data[headers[i]] = str(val)
                    text = "\n".join(f"{k}: {v}" for k, v in row_data.items() if v)
                    if text.strip():
                        docs.append(Document(page_content=text, metadata={"source": xlsx_file.name, "sheet": sheet_name}))
                        count += 1
                log.info(f"     -> {count} rows from sheet '{sheet_name}'")
            wb.close()
        except Exception as e:
            log.warning(f"     Failed to load {xlsx_file.name}: {e}")
    return docs


def load_all_documents(data_dir):
    """Load documents from all supported file types."""
    log.info(f"Scanning directory: {data_dir.absolute()}")
    all_docs = []
    all_docs.extend(load_pdfs(data_dir))
    all_docs.extend(load_csvs(data_dir))
    all_docs.extend(load_json_exercises(data_dir))
    all_docs.extend(load_xlsx(data_dir))
    log.info(f"Total documents loaded: {len(all_docs)}")
    return all_docs


def chunk_documents(docs):
    """Split documents into smaller chunks for embedding."""
    splitter = RecursiveCharacterTextSplitter(
        chunk_size=CHUNK_SIZE,
        chunk_overlap=CHUNK_OVERLAP,
        separators=["\n\n", "\n", ". ", " ", ""],
    )
    chunks = splitter.split_documents(docs)
    log.info(f"Created {len(chunks)} chunks from {len(docs)} documents")
    return chunks


def build_vectorstore(chunks, embeddings, force_rebuild=False):
    """Build or load a persistent ChromaDB vector store."""
    if not force_rebuild and CHROMA_DIR.exists() and any(CHROMA_DIR.iterdir()):
        log.info("ChromaDB already exists. Run with --rebuild to recreate.")
        return

    log.info("Building new ChromaDB vector store...")
    if CHROMA_DIR.exists():
        try:
            shutil.rmtree(CHROMA_DIR)
        except PermissionError:
            log.error(f"\n[LOCK ERROR] The folder {CHROMA_DIR.absolute()} is being used by another process.")
            log.error("Please STOP your Flask app (app.py) or any other gym tools first, then try again.\n")
            sys.exit(1)

    BATCH_SIZE = 1000
    for i in range(0, len(chunks), BATCH_SIZE):
        batch = chunks[i:i + BATCH_SIZE]
        log.info(f"  Embedding batch {i//BATCH_SIZE + 1}...")
        Chroma.from_documents(
            documents=batch,
            embedding=embeddings,
            persist_directory=str(CHROMA_DIR),
        )
    log.info("Vector database built successfully!")


def main():
    parser = argparse.ArgumentParser(description="Fitness Knowledge Base Vectorizer")
    parser.add_argument("--rebuild", action="store_true", help="Force rebuild the vector store")
    args = parser.parse_args()

    print("\n" + "=" * 60)
    print("  FITNESS KNOWLEDGE BASE VECTORIZER")
    print("=" * 60)

    # 1. Embeddings
    import torch
    embeddings = HuggingFaceEmbeddings(
        model_name=EMBEDDING_MODEL,
        model_kwargs={"device": "cuda" if torch.cuda.is_available() else "cpu"},
    )

    # 2. Loading
    docs = load_all_documents(DATA_DIR)
    if not docs:
        log.error(f"No documents found in {DATA_DIR.absolute()}. Check your dataset folder!")
        sys.exit(1)

    # 3. Chunking
    chunks = chunk_documents(docs)

    # 4. Building
    build_vectorstore(chunks, embeddings, force_rebuild=args.rebuild)


if __name__ == "__main__":
    main()
